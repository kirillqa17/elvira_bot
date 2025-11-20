import telebot
from telebot import types
import os
import threading
from dotenv import load_dotenv
import time
import openpyxl
from openpyxl.utils import get_column_letter
from database import * # Импортируем все, включая update_user_stage
from datetime import datetime
import pytz

load_dotenv()

TOKEN = str(os.getenv("BOT_TOKEN"))

bot = telebot.TeleBot(TOKEN)

# Хранение ответов пользователя
user_data = {}

# Хранение данных рассылки
broadcast_data = {}

with open("admins.txt", "r") as f:
    ADMINS = [int(line.strip()) for line in f if line.strip().isdigit()]

init_db()

# --- Админ-панель (без изменений в логике, кроме экспорта) ---
@bot.message_handler(commands=['admin'])
def admin_panel(message):
    if message.from_user.id not in ADMINS:
        return bot.send_message(message.chat.id, "⛔ У вас нет прав для входа в админ-панель")


    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("Реклама", callback_data="admin_ads"))
    markup.add(types.InlineKeyboardButton("База данных", callback_data="admin_db"))
    bot.send_message(message.chat.id, "⚙️ Админ-панель", reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data.startswith("admin"))
def handle_admin_menu(call):
    if call.from_user.id not in ADMINS:
        return bot.answer_callback_query(call.id, "Нет доступа")

    if call.data == "admin_main":
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("Реклама", callback_data="admin_ads"))
        markup.add(types.InlineKeyboardButton("База данных (экспорт XLSX)", callback_data="admin_db"))
        bot.edit_message_text("⚙️ Админ-панель", call.message.chat.id, call.message.message_id, reply_markup=markup)
        bot.answer_callback_query(call.id)
        return

    if call.data == "admin_ads":
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("Сгенерировать маркер", callback_data="admin_gen_marker"))
        markup.add(types.InlineKeyboardButton("Посмотреть все маркеры", callback_data="admin_view_markers"))
        markup.add(types.InlineKeyboardButton("Назад", callback_data="admin_main"))
        bot.edit_message_text("📢 Реклама:", call.message.chat.id, call.message.message_id, reply_markup=markup)



    elif call.data == "admin_db":
        bot.answer_callback_query(call.id, "Запускаю экспорт...")
        # Запускаем в отдельном потоке
        threading.Thread(target=generate_and_send_excel, args=(call.message.chat.id,), daemon=True).start()


    elif call.data == "admin_gen_marker":
        msg = bot.send_message(call.message.chat.id, "Введите название для маркера:")
        bot.register_next_step_handler(msg, process_marker_name)
        bot.answer_callback_query(call.id)


    elif call.data == "admin_view_markers":
        markers = get_markers()
        if not markers:
            bot.send_message(call.message.chat.id, "Маркеров пока нет.")
        else:
            lines = ["📋 Список маркеров:"]
            for mid, name, marker, created, users_total in markers:
                lines.append(f"➡️ {name}  |  `{marker}`  |  {users_total} юзеров  |  {created}")
            text = "\n".join(lines)
            bot.send_message(call.message.chat.id, text, parse_mode='Markdown')
        bot.answer_callback_query(call.id)


def process_marker_name(message):
    name = (message.text or "").strip()
    if not name:
        return bot.send_message(message.chat.id, "Название не может быть пустым. Повторите команду /admin → Реклама → Сгенерировать маркер")

    try:
        marker = create_marker(name)  # функция из database.py
        bot_username = bot.get_me().username or "<bot>"
        link = f"https://t.me/{bot_username}?start={marker}"

        # клавиатура после создания маркера
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("Посмотреть все маркеры", callback_data="admin_view_markers"))
        markup.add(types.InlineKeyboardButton("Назад в меню", callback_data="admin_main"))

        # Отправляем ПЛЕЙН-текст (без parse_mode), чтобы избежать ошибок парсинга
        text = f"✅ Маркер создан!\n\nНазвание: {name}\nКод: {marker}\nСсылка: {link}"
        bot.send_message(message.chat.id, text, reply_markup=markup)
    except Exception as e:
        bot.send_message(message.chat.id, f"Ошибка при создании маркера: {e}")


def generate_and_send_excel(admin_chat_id):
    """
    Генерирует Excel-файл с двумя листами (Users и Markers) и отправляет администратору.
    Работает в отдельном потоке.
    """
    timestamp = int(time.time())
    xlsx_name = f"export_{timestamp}.xlsx"
    try:
        # get_all_users() теперь вернет и current_stage
        users = get_all_users()
        markers = get_markers()

        # Создаём Excel
        wb = openpyxl.Workbook()

        # Лист Users
        ws_users = wb.active
        ws_users.title = "Users"
        # Добавляем новый заголовок
        headers_users = ['id', 'telegram_id', 'username', 'first_name', 'date_registered', 'ref_marker', 'current_stage']
        ws_users.append(headers_users)
        for row in users:
            ws_users.append(row)

        # Лист Markers
        ws_markers = wb.create_sheet(title="Markers")
        headers_markers = ['id', 'name', 'marker', 'created_at', 'users_total']
        ws_markers.append(headers_markers)
        for row in markers:
            ws_markers.append(row)

        # Автоширина колонок
        for ws in [ws_users, ws_markers]:
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[col_letter].width = adjusted_width

        # Сохраняем
        wb.save(xlsx_name)

        # Отправляем
        with open(xlsx_name, 'rb') as xf:
            bot.send_document(admin_chat_id, xf, caption="Экспорт базы данных (Users + Markers)")

    except Exception as e:
        try:
            bot.send_message(admin_chat_id, f"Произошла ошибка при экспорте базы: {e}")
        except Exception:
            pass
    finally:
        if os.path.exists(xlsx_name):
            try:
                os.remove(xlsx_name)
            except Exception:
                pass

# --- Основная логика бота с отслеживанием этапов ---

@bot.message_handler(commands=['start'])
def send_welcome(message):
    if message.chat.id in ADMINS:
        bot.send_message(message.chat.id, f"""Приветвую, {message.from_user.first_name}! Вы - администратор.
Для вызова меню управления отправьте /admin""")

    marker = None
    if message.text and message.text.startswith("/start "):
        marker = message.text.split(" ", 1)[1].strip()
    try:
        # Функция add_or_update_user уже установит этап 'start' для новых пользователей
        add_or_update_user(message.from_user, marker)
    except Exception as e:
        print("Ошибка при записи в БД:", e)

    video_path = 'media/video.mp4'
    with open(video_path, 'rb') as video_note:
        bot.send_video_note(message.chat.id, video_note)

    name = message.from_user.first_name
    greeting_text = f"""Привет, {name}! 👋🏻 Очень рада видеть тебя здесь!

Здесь всё по-настояшему просто, удобно и с заботой о тебе.

<b>Я помогу тебе мягко и уверенно двигаться к твоим целям, какие бы они ни были</b>:
🌸 снять отёки и вернуть ощущение лёгкости
🌸 выглядеть на 15 лет моложе без ботокса и хирургии
🌸 подтянуть живот и все тело
🌸 укрепить мышцы тазового дна
🌸 подтянуть овал лица и убрать второй подбородок
🌸 почувствовать больше энергии
"""
    photo_path = "media/intro.jpg"
    with open(photo_path, 'rb') as photo:
        bot.send_photo(
            chat_id=message.chat.id,
            photo=photo,
            caption=greeting_text, parse_mode='HTML'
        )
    # Обновляем этап
    update_user_stage(message.chat.id, "1_sent_welcome")

    threading.Timer(10, choose_branch, args=[message.chat.id]).start()

def choose_branch(chat_id):
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("Красивая осанка и здоровая спина", callback_data="branch_back"))
    markup.add(types.InlineKeyboardButton("Тело без отеков", callback_data="branch_body"))
    markup.add(types.InlineKeyboardButton("Молодое лицо и длинная шея", callback_data="branch_face"))

    bot.send_message(chat_id, "<b>Давай определим, какая тренировка нужна твоему телу прямо сейчас?</b>", reply_markup=markup, parse_mode='HTML')
    # Обновляем этап
    update_user_stage(chat_id, "2_sent_branch_choice")

@bot.callback_query_handler(func=lambda call: call.data.startswith("branch_"))
def second_stage(call):
    photo_path = 'media/elvira_photo.jpg'
    branch = call.data.split('branch_')[1]
    
    # Обновляем этап
    update_user_stage(call.message.chat.id, f"3_chose_branch_{branch}")
    
    with open(photo_path, 'rb') as photo:
        bot.send_photo(
            chat_id=call.message.chat.id,
            photo=photo,
            caption="""<b>Давай сделаю тебе приятно</b> 😏

Да-да, тебе не показалось))

Хочу поделиться с тобой пользой и показать, как избавиться от отеков навсегда, похудеть и даже помолодеть на 10-13 лет без косметологов и хирургов, потратив в день 10-15 минут 🚀

Для начала давай познакомимся! Меня зовут Эльвира Андриянова, и я эксперт по естественному омоложению, фейсфитнесу, похудению и тренер по осанке с 24х летним опытом!

<i>Мой путь начался с диагноза «порок сердца», операции, непонимания, как вернуть свое тело, здоровье, красоту 😅А сейчас мне 43 года, а выгляжу на 30; я полна сил и энергии; 2 года назад вышла замуж; не болею уже лет 10 и чувствую себя просто великолепно.</i>

А еще я знаю секретные способы, как забыть про отеки НАВСЕГДА, избавиться от лишнего веса, и даже в 40+ выглядеть на 10-13 лет моложе, быть сексуальной, энергичной и с невероятной ЖЕНСКОЙ ЭНЕРГИЕЙ 😍

<b>Так вот, к чему я?
А к тому, что я знаю все про молодое и здоровое тело! Поэтому в праве делиться этими знаниями с тобой</b> ❤️
""", parse_mode='HTML'
        )
    
    threading.Timer(30, before_free_complex, args=[call, branch]).start()

def clean_check_sub(chat_id, branch, timer = None):
    try:
        chat_member = bot.get_chat_member(chat_id=-1002039278578, user_id=chat_id)
        if chat_member.status in ['member', 'administrator', 'creator']:
            free_complex(chat_id, branch)
            update_user_stage(chat_id, "3.5_sent_subscription_prompt")
        else:
            if get_user_stage(chat_id).startswith("3_chose_branch_") and timer==10800:
                free_complex(chat_id, branch)
            elif get_user_stage(chat_id).startswith("3_chose_branch_") and timer==900:
                markup = types.InlineKeyboardMarkup()
                markup.add(types.InlineKeyboardButton("Подписаться", url="https://t.me/+p6N8QLUi2fo3YTc6"))
                markup.add(types.InlineKeyboardButton("Проверить подписку", callback_data=f"clean_check_sub_{branch}"))
                username = get_username(chat_id)
                bot.send_message(chat_id, f"{username}, не вижу твоей подписки, скорее подписывайся и смотри комплекс ❤️", reply_markup=markup)

    except Exception as e:
        print(f"Ошибка проверки подписки: {e}")



def before_free_complex(call, branch):
    text = """<b>Тук-тук, бежала к тебе со всех ног, чтобы сообщить, что тебя уже ждёт первый комплекс </b>

Но, прежде, подпишись на мой канал, чтобы мы обменялись взаимной энергией с тобой 😉

<i>Сразу после подписки, тебе придет комплекс </i> 🫶🏻
"""
    chat_id = call.message.chat.id
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("Подписаться", url="https://t.me/+p6N8QLUi2fo3YTc6"))
    markup.add(types.InlineKeyboardButton("Проверить подписку", callback_data=f"clean_check_sub_{branch}"))
    photo_path='media/running.jpg'
    with open(photo_path, 'rb') as photo:
        bot.send_photo(
            chat_id=chat_id,
            photo=photo,
            caption= text,
            parse_mode='HTML',
            reply_markup=markup
        )
    threading.Timer(900, clean_check_sub, args=[chat_id, branch, 900]).start()
    threading.Timer(10800, clean_check_sub, args=[chat_id, branch, 10800]).start()


def free_complex(chat_id, branch):
    if branch == "back":
        photo_path = 'media/free_complex_back.jpg'
        text = """<b>Сделай шаг навстречу...
красивому и молодому телу</b>

<b>Лови комплекс «Королевская осанка за 5 минут», сразу после которой ты почувствуешь:</b>
•легкость
•осанка выпрямиться
•больше свободы в движениях и прилив энергии
•подтянутый животик за счет ровной спины
"""
        yt_link = "https://youtu.be/JrohbzoOdpY"
        rt_link = "https://rutube.ru/video/private/b948ba4a8703db58f51cb47535de30c9/?p=bptaDMpy5tyuYJvxwPuqVw"
        vk_link = "https://vk.com/video-225608115_456239138?list=ln-JvKo7HzWLiZLH29jHz"

    elif branch == "face":
        photo_path = 'media/free_complex_face.jpg'
        text = """<b>Сделай шаг навстречу...
красивому и молодому лицу без брылей и морщин</b>

<b>Лови комплекс «Молодое и подтянутое лицо без брылей и морщин», сразу после которой ты почувствуешь:</b>
•расслабление мышц лица, ты почувствуешь «свободу» лица в прямом смысле
•овал лица подтянется
•заметно уменьшатся морщины и брыли
"""
        yt_link = "https://youtu.be/OreFR3YkOUM"
        rt_link = "https://rutube.ru/video/private/54d25449d8e9d87030df379e7814b2ee/?p=ouOBiOVJGbIDGNJ9JUlRfw"
        vk_link = "https://vk.com/video-225608115_456239137?list=ln-huwCY17lBjqCa7e8ex"

    elif branch == "body":
        photo_path = 'media/free_complex_body.jpg'
        text = """<b>Сделай шаг навстречу...
телу без отеков и лишней жидкости</b>

<b>Лови комплекс «-500 гр лишней жидкости сразу после тренировки», сразу после которой ты почувствуешь:</b>
•легкость
•уйдут -200-500 гр лишней жидкости
•больше свободы в движениях и прилив энергии
"""
        yt_link = "https://youtu.be/ItvDCRLp2iY"
        rt_link = "https://rutube.ru/video/private/a1120e480795632baed7c61121e0fda5/?p=DRXjSaCEkAveI7JHiX1XSw"
        vk_link = "https://vk.com/video-225608115_456239141?list=ln-WvrIAeAAqXHNdU6L1j"

    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("Смотреть на УouТube", url=yt_link))
    markup.add(types.InlineKeyboardButton("Смотреть на Rutube", url=rt_link))
    markup.add(types.InlineKeyboardButton("Смотреть на VK видео", url=vk_link))

    with open(photo_path, 'rb') as photo:
        bot.send_photo(
            chat_id=chat_id,
            photo=photo,
            caption= text,
            parse_mode='HTML',
            reply_markup=markup
        )
    # Обновляем этап
    update_user_stage(chat_id, "4_sent_first_complex")
    threading.Timer(300, after_free_complex, args=[chat_id, branch]).start()

def after_free_complex(chat_id, branch):
    if branch == "back":
        callback_data="sub_branch_back"
    elif branch == "face":
        callback_data="sub_branch_face"
    elif branch == "body":
        callback_data="sub_branch_body"

    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("Выполнила", callback_data=callback_data))
    markup.add(types.InlineKeyboardButton("Не выполнила", callback_data=callback_data))
    bot.send_message(chat_id, f"{bot.get_chat(chat_id).first_name}, ты выполнила комлпекс?", reply_markup=markup)
    # Обновляем этап
    update_user_stage(chat_id, "5_asked_about_completion")

@bot.callback_query_handler(func=lambda call: call.data.startswith("clean_check_sub_"))
def handle_clean_check_sub(call):
    branch = call.data.split('clean_check_sub_')[1]
    chat_id = call.message.chat.id
    try:
        chat_member = bot.get_chat_member(chat_id=-1002039278578, user_id=call.from_user.id)
        if chat_member.status in ['member', 'administrator', 'creator']:
            bot.answer_callback_query(call.id, "Спасибо за подписку! Сейчас подберу для тебя комплекс...")
            free_complex(chat_id, branch)
            update_user_stage(chat_id, "3.5_sent_subscription_prompt")
        else:
            bot.answer_callback_query(call.id, "❌ Кажется, ты еще не подписалась. Пожалуйста, подпишись и нажми снова", show_alert=True)
    except Exception as e:
        print(f"Ошибка проверки подписки: {e}")
        bot.answer_callback_query(call.id, "Произошла ошибка при проверке. Попробуй еще раз позже", show_alert=True)

@bot.callback_query_handler(func=lambda call: call.data.startswith("sub_branch_"))
def subscription_stage(call):
    branch = call.data.split('sub_branch_')[1]
    text = """<b>А теперь секретное задание для тебя </b>😎
Чтобы я могла дать тебе максимум пользы, подпишись на мой тг канал. 
В нём ты узнаешь все секреты как выглядишь моложе на 10-13 лет, забудешь про отеки и лишние кг

<b>А еще…стабильно будешь получать бесплатные комплексы для лица и тела, а также еженедельные ответы на вопросы конкретно по твоей ситуации</b>
"""
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("Подписаться", url="https://t.me/+p6N8QLUi2fo3YTc6"))
    markup.add(types.InlineKeyboardButton("Проверить подписку", callback_data=f"check_subscription_{branch}"))
    photo_path = "media/mem.jpg"
    with open(photo_path, 'rb') as photo:
        bot.send_photo(
            chat_id=call.message.chat.id,
            photo=photo,
            caption=text,
            parse_mode='HTML',
            reply_markup=markup
        )
    # Обновляем этап
    update_user_stage(call.message.chat.id, "6_sent_subscription_prompt")
    threading.Timer(86400, check_if_subed, args=[call.message.chat.id, call.from_user.id, branch]).start()

def check_if_subed(chat_id, user_id, branch):
    try:
        chat_member = bot.get_chat_member(chat_id=-1002039278578, user_id=user_id)
        if not chat_member.status in ['member', 'administrator', 'creator']:
            markup = types.InlineKeyboardMarkup()
            markup.add(types.InlineKeyboardButton("Разбери меня", url="https://t.me/+p6N8QLUi2fo3YTc6"))
            bot.send_message(chat_id, "Ты подписалась? В канале лично разбираю конкретные запросы каждой БЕСПЛАТНО ☺️", reply_markup=markup)
            # Обновляем этап
            update_user_stage(chat_id, "7_sent_subscription_reminder")
            threading.Timer(86400, second_complex, args=[chat_id, branch]).start()
    except Exception as e:
        print(f"Ошибка в check_if_subed (вероятно, пользователь заблокировал бота): {e}")

@bot.callback_query_handler(func=lambda call: call.data.startswith("check_subscription_"))
def check_subscription_callback(call):
    branch = call.data.split('check_subscription_')[1]
    chat_id = call.message.chat.id
    try:
        chat_member = bot.get_chat_member(chat_id=-1002039278578, user_id=call.from_user.id)
        if chat_member.status in ['member', 'administrator', 'creator']:
            bot.answer_callback_query(call.id, "Спасибо за подписку! Сейчас подберу для тебя комплекс...")
            # Обновляем этап
            update_user_stage(chat_id, "8a_subscription_confirmed")
            second_complex(chat_id, branch)
        else:
            bot.answer_callback_query(call.id, "❌ Кажется, ты еще не подписалась. Пожалуйста, подпишись и нажми снова", show_alert=True)
            # Обновляем этап
            update_user_stage(chat_id, "8b_subscription_failed")
    except Exception as e:
        print(f"Ошибка проверки подписки: {e}")
        bot.answer_callback_query(call.id, "Произошла ошибка при проверке. Попробуй еще раз позже", show_alert=True)


def second_complex(chat_id, branch):
    if branch == "back":
        photo_path = 'media/second_complex_back.jpg'
        text = f"""<b>{bot.get_chat(chat_id).first_name}, кажется мы не закончили….</b>

Ты тоже так считаешь…? Тогда тебя ждет следующий бесплатный комплекс «Здоровая спина и подтянутый живот» ❤️
"""
        yt_link = "https://youtu.be/LNOJyXNlej0"
        rt_link = "https://rutube.ru/video/private/d85f38329b99b6769ca0650bd045f258/?p=fGTXhZJ3D-pGHb7OUcphHg"
        vk_link = "https://vk.com/video-225608115_456239139?list=ln-8XwK7zhbPR4PuZbU3E"

    elif branch == "face":
        photo_path = 'media/second_complex_face.jpg'
        text = f"""<b>{bot.get_chat(chat_id).first_name}, кажется мы не закончили….</b>

Ты тоже так считаешь…? Тогда тебя ждет следующий бесплатный комплекс «Длинная и изящная шея, как у Нефертити» ❤️
"""
        yt_link = "https://youtu.be/Td53U_3TXY4"
        rt_link = "https://rutube.ru/video/private/5aabdbe378eca9683a9bfcf8f0cbd236/?p=6jlgx41csue5-s5z4fl_VA"
        vk_link = "https://vk.com/video-225608115_456239136?list=ln-MnMXCrWyKc4Trv2Qh0"

    elif branch == "body":
        photo_path = 'media/second_complex_body.jpg'
        text = f"""<b>{bot.get_chat(chat_id).first_name}, кажется мы не закончили….</b>

Ты тоже так считаешь…? Тогда тебя ждет следующий бесплатный комплекс «Лицо без отеков» ❤️
"""
        yt_link = "https://youtu.be/BTfh4rITS-A"
        rt_link = "https://rutube.ru/video/private/eb97aaaf0c65704d843092c4c4d04e61/?p=V0nNuWhJnEq22boDgCGLqQ"
        vk_link = "https://vk.com/video-225608115_456239140?list=ln-E87Sspq9OKLTXL9bK2"

    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("Смотреть на УouТube", url=yt_link))
    markup.add(types.InlineKeyboardButton("Смотреть на Rutube", url=rt_link))
    markup.add(types.InlineKeyboardButton("Смотреть на VK видео", url=vk_link))

    with open(photo_path, 'rb') as photo:
        bot.send_photo(
            chat_id=chat_id,
            photo=photo,
            caption= text,
            parse_mode='HTML',
            reply_markup=markup
        )
    # Обновляем этап
    update_user_stage(chat_id, "9_sent_second_complex")
    threading.Timer(86400, send_day_after_message, args=[chat_id]).start()


def send_day_after_message(chat_id):
    try:
        name = bot.get_chat(chat_id).first_name
    except Exception as e:
        print(f"Не удалось получить чат {chat_id}: {e}")
        return 

    success_story = f"""<b>Нам нужно серьезно поговорить!</b>

От «нелюбви к себе» до «я вновь полюбила себя и почувствовала женщиной» 🔥
Знакомьтесь, на фото - Ира, 39 лет

❌<b>Ира пришла ко мне с</b>: лишним весом, отечностью, двойным подбородком, болью в шее и спине, целлюлитом, сутулостью и ощущением, что уже ничего не изменить…

✅ <b>Спустя 4 недели:</b> ушли отеки и лишние килограммы, улучшилась осанка и перестала болеть спина, ушла отечность с лица, научилась нравиться себе и раскрыла себя как женщину.

<b>Что помогло нам сделать такой результат:</b>
-10-15ти минутные тренировки в день ✅
-Питание без диет и жестких запретов ✅
-Танцы, которые приносят удовольствие ✅
-Проработка сексуальности и женственности ✅

<b>И так может каждая, если выбрана верная система действий!</b>
"""
    media = [
        types.InputMediaVideo(open('media/ira.mp4','rb')),
        types.InputMediaPhoto(open('media/ira_photo.jpg','rb'))
    ]
    bot.send_media_group(chat_id, media)
    # Обновляем этап
    update_user_stage(chat_id, "10_sent_ira_story")
    threading.Timer(86400, send_follow_up_message, args=[chat_id]).start()
    bot.send_message(chat_id, success_story, parse_mode='HTML')

def send_follow_up_message(chat_id):
    try:
        name = bot.get_chat(chat_id).first_name
    except Exception as e:
        print(f"Не удалось получить чат {chat_id}: {e}")
        return 

    follow_up_text = f"""{name}, ты так помолодела и постройнела 🔥

<b>Хочешь получать такие же комплименты?</b>

Это секрет работы с осанкой и отеками, с мышцами лица
а не диет и жестких тренировок, походов к косметологу 

У девушек подтянулся овал лица, вес пошел вниз, ушел второй подбородок 

<b>При этом они тренировалась на лайте:</b>
- 3-4 раза в неделю дома
- делали 15-30 минутные комплексы и омолаживались даже ночью 

🏆 <b>Итог: ушли отеки, подтянулся живот, ушел второй подбородок, осанка стала КОРОЛЕВСКОЙ</b>

Доказательства? (В фото выше 👆🏻)

Ставь ❤️ если переживаешь, что не сможешь сделать тело мечты и помолодеть на 10-13 лет
"""

    media = [
        types.InputMediaPhoto(open('media/results1.jpg', 'rb')),
        types.InputMediaPhoto(open('media/results3.jpg', 'rb'))
    ]
    bot.send_media_group(chat_id, media)
    markup = types.InlineKeyboardMarkup()
    btn_join = types.InlineKeyboardButton(
        text="❤️",
        url="https://t.me/Elvira_ELAN"
    )
    markup.add(btn_join)
    bot.send_message(chat_id, follow_up_text, parse_mode='HTML', reply_markup=markup)
    # Обновляем этап
    update_user_stage(chat_id, "11_sent_results_followup")
    threading.Timer(86400, send_final_pitch, args=[chat_id]).start()


def send_final_pitch(chat_id):
    final_text = f"""❌ <b>Это вообще нормально? </b>

Я и мои клиентки выглядим на 13 лет моложе без косметологов и хирургов, без потраченных 100.000+ на все эти инъекции и процедуры!

<b>Знакомый ужас?</b>
➡️ Вложила в свое лицо и тело больше 100k — а результата ноль.
➡️ Делаешь маски и мажешься кремами, а морщины и отечность на месте.
➡️ Уже ненавидишь себя и свое отражение в зеркале.
➡️ Мечтаешь проснуться одним днем моложе на 10-13 лет.

А тем временем, косметологи и хирурги рассказывают, сколько нужно сделать различных “безопасных” процедур и операций, чтобы выглядеть хорошо.

<u>Я тоже верила в этот бред. Пока не включила голову и не нашла способ, который работает в миллион раз лучше и дешевле.</u>

И сейчас я и мои девочки, которые еще вчера паниковали из-за своего возраста, просто смеются над тем, что когда-то сомневались):

<b>Это не про «просто повезло» — это система. Это про то, что:</b>
👍 Не нужно ходить к косметологам и отдавать последние деньги 
👍 Не нужно класть свое тело и лицо под нож
👍 Не нужно убиваться в спортзале каждый день

Звучит как сказка? Но нет, это реальность 🔥
Готова раскрыть систему и показать, как стать молодой, красивой и здоровой даже после 40 ↴

<b>Ставь «+», и я проведу для тебя БЕСПЛАТНУЮ консультацию, на которой разберу конкретно твою ситуацию и дам четкие шаги, как сделать тело и лицо мечты даже после 40 (другие продают это дорого, а я делаю бесплатно)</b>
"""
    media = [
        types.InputMediaPhoto(open('media/final1.jpg', 'rb')),
        types.InputMediaPhoto(open('media/final2.jpg', 'rb')),
        types.InputMediaPhoto(open('media/final3.jpg', 'rb'))     
    ] 
    bot.send_media_group(chat_id, media)
    markup = types.InlineKeyboardMarkup()
    btn_join = types.InlineKeyboardButton(
        text="+",
        url="https://t.me/Elvira_ELAN"
    )
    markup.add(btn_join)
    bot.send_message(chat_id, final_text, reply_markup=markup, parse_mode='HTML')
    update_user_stage(chat_id, "12_sent_final_pitch")


# --- Рассылка для администраторов ---

@bot.message_handler(commands=['send'])
def send_broadcast_command(message):
    """Команда для запуска рассылки (только для админов)"""
    if message.from_user.id not in ADMINS:
        return bot.send_message(message.chat.id, "⛔ У вас нет прав для создания рассылки")

    bot.send_message(message.chat.id, "📢 Создание рассылки\n\nОтправьте сообщение, которое будет разослано всем пользователям в базе данных.\n\n<i>Сообщение может содержать текст, фото, видео и т.д.</i>", parse_mode='HTML')
    bot.register_next_step_handler(message, receive_broadcast_message)


def receive_broadcast_message(message):
    """Получение сообщения для рассылки"""
    if message.from_user.id not in ADMINS:
        return

    # Сохраняем сообщение
    broadcast_data[message.from_user.id] = {
        'message': message
    }

    # Отправляем предпросмотр
    bot.send_message(message.chat.id, "📋 <b>Предпросмотр сообщения для рассылки:</b>", parse_mode='HTML')

    try:
        # Копируем сообщение админу как предпросмотр
        if message.content_type == 'text':
            if message.entities:
                bot.send_message(message.chat.id, message.text, entities=message.entities)
            else:
                bot.send_message(message.chat.id, message.text)

        elif message.content_type == 'photo':
            if message.caption_entities:
                bot.send_photo(message.chat.id, message.photo[-1].file_id, caption=message.caption, caption_entities=message.caption_entities)
            else:
                bot.send_photo(message.chat.id, message.photo[-1].file_id, caption=message.caption)

        elif message.content_type == 'video':
            if message.caption_entities:
                bot.send_video(message.chat.id, message.video.file_id, caption=message.caption, caption_entities=message.caption_entities)
            else:
                bot.send_video(message.chat.id, message.video.file_id, caption=message.caption)

        elif message.content_type == 'document':
            if message.caption_entities:
                bot.send_document(message.chat.id, message.document.file_id, caption=message.caption, caption_entities=message.caption_entities)
            else:
                bot.send_document(message.chat.id, message.document.file_id, caption=message.caption)

        elif message.content_type == 'audio':
            if message.caption_entities:
                bot.send_audio(message.chat.id, message.audio.file_id, caption=message.caption, caption_entities=message.caption_entities)
            else:
                bot.send_audio(message.chat.id, message.audio.file_id, caption=message.caption)

        elif message.content_type == 'voice':
            bot.send_voice(message.chat.id, message.voice.file_id)

        elif message.content_type == 'video_note':
            bot.send_video_note(message.chat.id, message.video_note.file_id)

        elif message.content_type == 'sticker':
            bot.send_sticker(message.chat.id, message.sticker.file_id)

        elif message.content_type == 'animation':
            if message.caption_entities:
                bot.send_animation(message.chat.id, message.animation.file_id, caption=message.caption, caption_entities=message.caption_entities)
            else:
                bot.send_animation(message.chat.id, message.animation.file_id, caption=message.caption)
        else:
            # Для других типов просто копируем сообщение
            bot.copy_message(message.chat.id, message.chat.id, message.message_id)
    except Exception as e:
        bot.send_message(message.chat.id, f"⚠️ Ошибка при создании предпросмотра: {e}")

    bot.send_message(message.chat.id,
                    "✅ Сообщение для рассылки получено!\n\n"
                    "Теперь укажите дату и время начала рассылки по Москве в формате:\n"
                    "<code>ДД.ММ.ГГГГ ЧЧ:ММ</code>\n\n"
                    "Например: <code>25.10.2025 14:30</code>\n\n"
                    "<i>Или отправьте 'сейчас' для немедленной рассылки</i>",
                    parse_mode='HTML')
    bot.register_next_step_handler(message, receive_broadcast_datetime)


def receive_broadcast_datetime(message):
    """Получение даты и времени для рассылки"""
    if message.from_user.id not in ADMINS:
        return

    if message.from_user.id not in broadcast_data:
        return bot.send_message(message.chat.id, "❌ Ошибка: сообщение для рассылки не найдено. Начните заново с /send")

    datetime_text = message.text.strip().lower()

    # Московский часовой пояс
    moscow_tz = pytz.timezone('Europe/Moscow')

    # Если пользователь хочет отправить сейчас
    if datetime_text == 'сейчас':
        schedule_time_moscow = datetime.now(moscow_tz)
        delay_seconds = 0
    else:
        # Парсим дату и время
        try:
            # Ожидаем формат: ДД.ММ.ГГГГ ЧЧ:ММ
            schedule_time_naive = datetime.strptime(datetime_text, '%d.%m.%Y %H:%M')
            # Делаем datetime aware в московском часовом поясе
            schedule_time_moscow = moscow_tz.localize(schedule_time_naive)

            # Проверяем, что время в будущем
            now_moscow = datetime.now(moscow_tz)
            if schedule_time_moscow <= now_moscow:
                return bot.send_message(message.chat.id,
                                       "❌ Указанное время уже прошло. Пожалуйста, укажите время в будущем.\n"
                                       "Отправьте дату и время снова в формате: <code>ДД.ММ.ГГГГ ЧЧ:ММ</code>",
                                       parse_mode='HTML')

            # Вычисляем задержку в секундах
            delay_seconds = (schedule_time_moscow - now_moscow).total_seconds()

        except ValueError:
            return bot.send_message(message.chat.id,
                                   "❌ Неверный формат даты и времени.\n"
                                   "Пожалуйста, используйте формат: <code>ДД.ММ.ГГГГ ЧЧ:ММ</code>\n"
                                   "Например: <code>25.10.2025 14:30</code>",
                                   parse_mode='HTML')

    # Получаем сообщение для рассылки
    broadcast_message = broadcast_data[message.from_user.id]['message']

    # Планируем рассылку
    if delay_seconds == 0:
        bot.send_message(message.chat.id, "✅ Рассылка запущена немедленно!")
        threading.Thread(target=execute_broadcast, args=(broadcast_message,), daemon=True).start()
    else:
        formatted_time = schedule_time_moscow.strftime('%d.%m.%Y %H:%M')
        bot.send_message(message.chat.id,
                        f"✅ Рассылка запланирована на <b>{formatted_time}</b> (по Москве)\n\n"
                        f"До начала рассылки: {int(delay_seconds // 3600)} ч. {int((delay_seconds % 3600) // 60)} мин.",
                        parse_mode='HTML')
        threading.Timer(delay_seconds, execute_broadcast, args=(broadcast_message,)).start()

    # Очищаем данные
    del broadcast_data[message.from_user.id]


def execute_broadcast(message):
    """Выполняет рассылку всем пользователям"""
    users = get_all_users()
    total_users = len(users)
    success_count = 0
    failed_count = 0

    print(f"[BROADCAST] Начало рассылки для {total_users} пользователей")

    for user in users:
        telegram_id = user[1]  # telegram_id находится во второй колонке

        try:
            # Определяем тип сообщения и пересылаем его
            if message.content_type == 'text':
                # Для текстовых сообщений проверяем наличие entities
                if message.entities:
                    bot.send_message(telegram_id, message.text, entities=message.entities)
                else:
                    bot.send_message(telegram_id, message.text)

            elif message.content_type == 'photo':
                # Для фото с подписью
                if message.caption_entities:
                    bot.send_photo(telegram_id, message.photo[-1].file_id, caption=message.caption, caption_entities=message.caption_entities)
                else:
                    bot.send_photo(telegram_id, message.photo[-1].file_id, caption=message.caption)

            elif message.content_type == 'video':
                if message.caption_entities:
                    bot.send_video(telegram_id, message.video.file_id, caption=message.caption, caption_entities=message.caption_entities)
                else:
                    bot.send_video(telegram_id, message.video.file_id, caption=message.caption)

            elif message.content_type == 'document':
                if message.caption_entities:
                    bot.send_document(telegram_id, message.document.file_id, caption=message.caption, caption_entities=message.caption_entities)
                else:
                    bot.send_document(telegram_id, message.document.file_id, caption=message.caption)

            elif message.content_type == 'audio':
                if message.caption_entities:
                    bot.send_audio(telegram_id, message.audio.file_id, caption=message.caption, caption_entities=message.caption_entities)
                else:
                    bot.send_audio(telegram_id, message.audio.file_id, caption=message.caption)

            elif message.content_type == 'voice':
                if message.caption_entities:
                    bot.send_voice(telegram_id, message.voice.file_id, caption=message.caption, caption_entities=message.caption_entities)
                else:
                    bot.send_voice(telegram_id, message.voice.file_id, caption=message.caption)

            elif message.content_type == 'video_note':
                bot.send_video_note(telegram_id, message.video_note.file_id)

            elif message.content_type == 'sticker':
                bot.send_sticker(telegram_id, message.sticker.file_id)

            elif message.content_type == 'animation':
                if message.caption_entities:
                    bot.send_animation(telegram_id, message.animation.file_id, caption=message.caption, caption_entities=message.caption_entities)
                else:
                    bot.send_animation(telegram_id, message.animation.file_id, caption=message.caption)
            else:
                # Для других типов просто копируем сообщение
                bot.copy_message(telegram_id, message.chat.id, message.message_id)

            success_count += 1
            time.sleep(0.05)  # Небольшая задержка между отправками для избежания лимитов API

        except Exception as e:
            failed_count += 1
            print(f"[BROADCAST] Ошибка отправки пользователю {telegram_id}: {e}")

    # Отправляем отчёт админу
    report = (f"📊 Рассылка завершена!\n\n"
             f"Всего пользователей: {total_users}\n"
             f"Успешно отправлено: {success_count}\n"
             f"Ошибок: {failed_count}")

    for admin_id in ADMINS:
        try:
            bot.send_message(admin_id, report)
        except:
            pass

    print(f"[BROADCAST] Завершено. Успешно: {success_count}, Ошибок: {failed_count}")


if __name__ == '__main__':
    bot.polling(none_stop=True)